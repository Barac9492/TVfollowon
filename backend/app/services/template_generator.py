from __future__ import annotations
"""Generate growth metrics Excel templates."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GROWTH_HEADERS = [
    ("company_id", 12, "회사 ID (필수)"),
    ("company_name", 28, "회사명 (확인용)"),
    ("metric_date", 14, "측정일 (YYYY-MM-DD, 필수)"),
    ("monthly_revenue", 16, "월매출"),
    ("revenue_currency", 10, "통화 (KRW/USD)"),
    ("arr", 14, "ARR"),
    ("mrr", 14, "MRR"),
    ("revenue_at_first_meeting", 20, "첫 미팅 시 매출"),
    ("mrr_growth_rate_pct", 18, "MRR 성장률 (%, MoM)"),
    ("monthly_burn", 14, "월 번레이트"),
    ("cash_on_hand", 14, "보유 현금"),
    ("runway_months", 14, "런웨이 (개월)"),
    ("headcount", 10, "인원수"),
    ("key_metric_value", 16, "핵심 KPI 값"),
    ("key_metric_name", 16, "핵심 KPI 이름"),
    ("last_funding_date", 16, "최근 펀딩일"),
    ("last_funding_amount", 18, "최근 펀딩 금액"),
    ("last_funding_round", 16, "최근 라운드"),
    ("paying_customers", 14, "유료 고객수"),
    ("ndr_pct", 12, "NDR (%)"),
    ("notes", 30, "메모"),
]


def generate_growth_template(companies=None) -> bytes:
    """Generate a growth metrics Excel template.

    Args:
        companies: Optional list of (id, name) tuples to pre-fill.
    Returns:
        Excel file as bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Growth Metrics"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    desc_font = Font(italic=True, color="64748B", size=9)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Row 1: Headers
    for col, (name, width, _desc) in enumerate(GROWTH_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Row 2: Descriptions (Korean)
    for col, (_name, _width, desc) in enumerate(GROWTH_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=desc)
        cell.font = desc_font
        cell.border = thin_border

    # Pre-fill company IDs if provided
    if companies:
        for i, (cid, cname) in enumerate(companies, 3):
            ws.cell(row=i, column=1, value=cid).border = thin_border
            ws.cell(row=i, column=2, value=cname).border = thin_border

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
