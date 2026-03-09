from __future__ import annotations
"""Excel parser for portfolio data and comments.

Handles headerless .xlsx files by detecting column types from data patterns.
"""
import openpyxl
from datetime import datetime
from pathlib import Path


def parse_portfolio_excel(file_path: str | Path) -> list[dict]:
    """Parse portfolio Excel file (File 1 format).

    Expected columns (no headers):
    0: id (string), 1: status, 2: batch, 3: company_name,
    4: representative_name, 5: current_valuation, 6: current_currency, 7: current_stage
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue

        # Skip header row if present
        first_val = str(row[0]).strip().lower()
        if first_val == "id":
            continue

        try:
            valuation = _parse_number(row[5]) if len(row) > 5 else None
            rows.append({
                "id": str(row[0]).strip(),
                "status": str(row[1]).strip() if row[1] else None,
                "batch": str(row[2]).strip() if row[2] else None,
                "company_name": str(row[3]).strip() if row[3] else None,
                "representative_name": str(row[4]).strip() if len(row) > 4 and row[4] else None,
                "current_valuation": valuation,
                "current_currency": str(row[6]).strip() if len(row) > 6 and row[6] else "KRW",
                "current_stage": str(row[7]).strip().lower() if len(row) > 7 and row[7] else "none",
            })
        except (IndexError, ValueError, TypeError):
            continue

    wb.close()
    return rows


def parse_comments_excel(file_path: str | Path) -> list[dict]:
    """Parse comments Excel file (File 2 format).

    Expected columns (no headers):
    0: comment, 1: company_name, 2: id (FK to company)
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue

        first_val = str(row[0]).strip().lower()
        if first_val == "comment":
            continue

        try:
            rows.append({
                "comment_text": str(row[0]).strip(),
                "company_name": str(row[1]).strip() if len(row) > 1 and row[1] else None,
                "company_id": str(row[2]).strip() if len(row) > 2 and row[2] else None,
            })
        except (IndexError, ValueError, TypeError):
            continue

    wb.close()
    return rows


def parse_growth_excel(file_path: str | Path) -> list[dict]:
    """Parse growth metrics Excel file (header-based).

    Expected headers: company_id, company_name, metric_date, monthly_revenue,
    revenue_currency, arr, mrr, revenue_at_first_meeting, mrr_growth_rate_pct,
    monthly_burn, cash_on_hand, runway_months, headcount, key_metric_value,
    key_metric_name, last_funding_date, last_funding_amount, last_funding_round,
    paying_customers, ndr_pct, notes
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    header_map = {}
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] is None:
            continue

        # First non-empty row = headers
        if not header_map:
            header_map = {str(cell).strip().lower(): i for i, cell in enumerate(row) if cell}
            if "company_id" not in header_map:
                # Try without header (fallback)
                wb.close()
                return []
            continue

        try:
            cid = str(row[header_map["company_id"]]).strip() if header_map.get("company_id") is not None else None
            if not cid:
                continue

            def _get(key):
                idx = header_map.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            metric_date = _get("metric_date")
            last_funding = _get("last_funding_date")

            rows.append({
                "company_id": cid,
                "company_name": str(_get("company_name")).strip() if _get("company_name") else None,
                "metric_date": _parse_date(metric_date),
                "monthly_revenue": _parse_number(_get("monthly_revenue")),
                "revenue_currency": str(_get("revenue_currency")).strip() if _get("revenue_currency") else "KRW",
                "arr": _parse_number(_get("arr")),
                "mrr": _parse_number(_get("mrr")),
                "revenue_at_first_meeting": _parse_number(_get("revenue_at_first_meeting")),
                "mrr_growth_rate_pct": _parse_number(_get("mrr_growth_rate_pct")),
                "monthly_burn": _parse_number(_get("monthly_burn")),
                "cash_on_hand": _parse_number(_get("cash_on_hand")),
                "runway_months": _parse_number(_get("runway_months")),
                "headcount": int(_parse_number(_get("headcount"))) if _parse_number(_get("headcount")) is not None else None,
                "key_metric_value": _parse_number(_get("key_metric_value")),
                "key_metric_name": str(_get("key_metric_name")).strip() if _get("key_metric_name") else None,
                "last_funding_date": _parse_date(last_funding),
                "last_funding_amount": _parse_number(_get("last_funding_amount")),
                "last_funding_round": str(_get("last_funding_round")).strip() if _get("last_funding_round") else None,
                "paying_customers": int(_parse_number(_get("paying_customers"))) if _parse_number(_get("paying_customers")) is not None else None,
                "ndr_pct": _parse_number(_get("ndr_pct")),
                "notes": str(_get("notes")).strip() if _get("notes") else None,
            })
        except (IndexError, ValueError, TypeError):
            continue

    wb.close()
    return rows


def _parse_date(value):
    """Parse date from Excel cell (datetime or string)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
    except Exception:
        pass
    return None


def _parse_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(",", "").strip()
        return float(cleaned)
    except (ValueError, TypeError):
        return None
